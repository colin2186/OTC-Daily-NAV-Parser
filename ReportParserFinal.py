import pandas as pd
import glob
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np


# Step 1: Dynamically find all matching files
file_path = r'C:\Users\cdunne\Documents\ASGARD'
file_pattern = 'ASGARD_OTCDerivativesReport-*.csv'  # Modified to match any date
matching_files = glob.glob(os.path.join(file_path, file_pattern))

if len(matching_files) == 0:
    print("No matching files found. Please check the directory and file pattern.")
    exit()

print(f"Found {len(matching_files)} files to process")


# Function to extract date from filename
def extract_date_from_filename(filename):
    # Extract date from filename like ASGARD_OTCDerivativesReport-20250129
    return filename.split('-')[-1].split('.')[0]


# Initialize an empty list to store DataFrames
all_data = []

# Process each file
for file_to_read in matching_files:
    print(f"Reading file: {file_to_read}")

    # Read the CSV and set the correct header row
    data = pd.read_csv(file_to_read, header=13)
    data.columns = data.columns.str.strip()

    # Add Report Date column based on filename
    report_date = extract_date_from_filename(file_to_read)
    data['Report Date'] = report_date

    # Rename specific columns
    data = data.rename(columns={
        'Unnamed: 0': 'Trade ID 1',
        'Unnamed: 1': 'Trade ID 2',
        'Unnamed: 5': 'Product Sub Type'
    })

    all_data.append(data)

# Combine all DataFrames
data = pd.concat(all_data, ignore_index=True)

# Remove rows where 'Trade ID 1' is missing (i.e., remove totals row)
data = data.dropna(subset=['Trade ID 1'])




# Step 4: Ask user for the NAV
try:
    nav = float(input("Please add ME NAV (e.g., 439607201): "))
except ValueError:
    print("Invalid NAV entered. Please enter a numeric value.")
    exit()

# Step 5: Map Excel-like column references (A, B, F, etc.) to actual column names
excel_columns = {
    'A': data.columns[0],  # Replace with the actual column index for A
    'B': data.columns[1],  # Column B
    'F': data.columns[5],  # Column F
    'V': data.columns[21], # Column V
    'W': data.columns[22], # Column W
    'X': data.columns[23], # Column X
    'AZ': data.columns[51], # Column AZ
    'BA': data.columns[52], # Column BA
    'BB': data.columns[53], # Column BB
    'BC': data.columns[54], # Column BC
    'BD': data.columns[55], # Column BD
    'BE': data.columns[56], # Column BE
    'BF': data.columns[57], # Column BF
    'BG': data.columns[58], # Column BG
    'AG': data.columns[32], # Column AG
    'BO': data.columns[66] # Column BO
}

# Step 6: Convert numeric columns to proper numeric format
numeric_columns = [excel_columns['W'], excel_columns['AG'], excel_columns['X'], excel_columns['V']]
for col in numeric_columns:
    data[col] = pd.to_numeric(data[col].replace(',', '', regex=True), errors='coerce')

# Step 7: Filter the DataFrame using the mapped columns
selected_columns = ['A', 'B', 'F', 'V','W','X', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'AG', 'BO']
df = data[[excel_columns[col] for col in selected_columns]].copy()

# Convert Final Source Load Time to DDMMYYYY format
df['Final Source Load Time'] = pd.to_datetime(df['Final Source Load Time']).dt.strftime('%d%m%Y')


def get_index(row):
    # Function to check if a value is a pure number (including negative numbers and percentages)
    def is_numeric_rate(value):
        if not isinstance(value, str):
            return True  # Already a number (float/int)
        # Remove '%', spaces, and check if it can be converted to a float
        cleaned = value.replace('%', '').replace(' ', '')
        return cleaned.replace('.', '', 1).replace('-', '', 1).isdigit()

    rec_rate = row['Rec Rate']
    pay_rate = row['Pay Rate']

    # Check if Rec Rate is NOT numeric, meaning it's likely an index
    if not is_numeric_rate(rec_rate):
        return rec_rate  # Rec Rate is a text (e.g., an index like "DKKCIBOR6M")
    # If Rec Rate is numeric, check Pay Rate
    elif not is_numeric_rate(pay_rate):
        return pay_rate  # Pay Rate is- a text (index)
    # If both are numeric, return None
    return None

# Apply the function to populate the Index column
df['Index'] = df.apply(get_index, axis=1)

# Optional: Verify the Index column doesn't contain any numeric values
numeric_indices = df['Index'].str.contains(r'^[\d\.]+%?$', na=False)
if numeric_indices.any():
    print("Warning: Some numeric values found in Index column")
    print(df[numeric_indices][['Rec Rate', 'Pay Rate', 'Index']])

# Step 8: Add new columns for tolerance checks
df['NAV Break (BPs)'] = (df[excel_columns['W']] / nav) * 10000
df['Sensitivity Break (BPs)'] = df[excel_columns['W']] / df[excel_columns['AG']]

# Step 9: Add Sensitivity Diff Check and NAV Break Check columns
df['Sensitivity Diff Check (BPs)'] = df[excel_columns['V']] - df['Sensitivity Break (BPs)']
df['NAV Break Check (BPs)'] = df[excel_columns['X']] - df['NAV Break (BPs)']

# Step 10: Round specific columns to 2 decimal places
columns_to_round = ['NAV Break (BPs)', 'Sensitivity Break (BPs)',
                    'Sensitivity Diff Check (BPs)', 'NAV Break Check (BPs)']
df[columns_to_round] = df[columns_to_round].round(2)

# Step 11: Set up new columns
df['Sensitivity Breach'] = None
df['Tolerance Breach'] = None

# Step 12: Ask the user which client they are analyzing
client = input("Please enter the client you are analyzing (e.g., ASGARD): ").strip()

if client.upper() == "ASGARD":
    # Apply conditions for Tolerance Breach
    df['Tolerance Breach'] = df['NAV Break (BPs)'].abs() > 1

# Define the currency list
ccy_list = ['USD', 'CAD', 'JPY', 'AUD', 'NZD', 'GBP', 'EUR', 'CHF', 'SEK', 'NOK']

# Apply Sensitivity Breach logic
df['Sensitivity Breach'] = df.apply(
    lambda row: (
        "TRUE" if (
            # For currencies in the list
            row['Ccy'] in ccy_list and
            (
                (row['Product Sub Type'] == "Plain Vanilla" and abs(row['Sensitivity Break (BPs)']) > 2) or
                (row['Product Sub Type'] == "OIS" and abs(row['Sensitivity Break (BPs)']) > 1) or
                (row['Product Sub Type'] == "MTM Cross Currency Swap" and abs(row['Sensitivity Break (BPs)']) > 4)
            )
        ) else (
            # For currencies outside the list
            "TRUE" if (
                row['Ccy'] not in ccy_list and
                (
                    (row['Product Sub Type'] == "Plain Vanilla" and abs(row['Sensitivity Break (BPs)']) > 5) or
                    (row['Product Sub Type'] == "OIS" and abs(row['Sensitivity Break (BPs)']) > 8) or
                    (row['Product Sub Type'] == "MTM Cross Currency Swap" and abs(row['Sensitivity Break (BPs)']) > 9)
                )
            ) else "FALSE"
        )
    ), axis=1
)

# Step 13: Save the updated DataFrame to an Excel file
output_file = os.path.join(file_path, f'Processed_ASGARD_Report_with_Breaches.xlsx')
df.to_excel(output_file, index=False, engine='openpyxl', sheet_name="Processed Report")
print(f"Processed data with breaches saved to: {output_file}")

# Step 14: Apply conditional formatting for TRUE/FALSE
wb = load_workbook(output_file)

ws = wb["Processed Report"]
# Ensure both columns are in string format for consistent conditional formatting
df['Sensitivity Breach'] = df['Sensitivity Breach'].astype(str)
df['Tolerance Breach'] = df['Tolerance Breach'].astype(str)

# Identify columns for conditional formatting
columns_to_format = ["Sensitivity Breach", "Tolerance Breach"]
true_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for TRUE
false_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for FALSE

# Apply conditional formatting for both columns
for col in columns_to_format:
    if col in df.columns:  # Ensure the column exists
        col_idx = df.columns.get_loc(col) + 1  # Get column index (1-based for Excel)
        for row in range(2, len(df) + 2):  # Start from row 2 (skip header)
            cell = ws.cell(row=row, column=col_idx)
            if str(cell.value).upper() == "TRUE":
                cell.fill = true_fill
            elif str(cell.value).upper() == "FALSE":
                cell.fill = false_fill
    else:
        print(f"Warning: Column '{col}' not found in DataFrame. Skipping conditional formatting.")

# Save the workbook with formatting
wb.save(output_file)
print(f"Conditional formatting applied and saved to: {output_file}")


# Ensure Tolerance Breach is a boolean before plotting
df['Tolerance Breach'] = df['Tolerance Breach'].str.upper() == "TRUE"

# Step 14: Visualization - Breaches grouped by Product Type and Ccy
df['Product_Ccy'] = df['Product Sub Type'] + "_" + df['Ccy']  # Combine Product Type and Ccy

# Count Sensitivity Breach (TRUE) grouped by Product_Ccy
sensitivity_breach_counts = df[df['Sensitivity Breach'] == "TRUE"]['Product_Ccy'].value_counts()

# Count Tolerance Breach (TRUE) grouped by Product_Ccy
tolerance_breach_counts = df[df['Tolerance Breach'] == True]['Product_Ccy'].value_counts()

# Align indices of both counts (fill missing values with 0)
all_product_ccy = sensitivity_breach_counts.index.union(tolerance_breach_counts.index)
sensitivity_breach_counts = sensitivity_breach_counts.reindex(all_product_ccy, fill_value=0)
tolerance_breach_counts = tolerance_breach_counts.reindex(all_product_ccy, fill_value=0)

# Plotting both charts stacked
fig, axes = plt.subplots(3, 1, figsize=(14, 16), sharex=True)

# Function to add values inside bars
def add_bar_labels(ax):
    for bar in ax.patches:
        height = bar.get_height()
        if height > 0:
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                height * 0.5,  # Position inside the bar
                str(int(height)),
                ha='center',
                va='center',  # Centered inside the bar
                fontsize=12,
                fontweight='bold',
                color='white'  # White text for contrast
            )

# Sensitivity Breach Chart
bars1 = axes[0].bar(
    sensitivity_breach_counts.index,
    sensitivity_breach_counts,
    color='skyblue',
    edgecolor='black'  # Solid border
)
axes[0].set_title("Count of Sensitivity Breaches by Product Type and Ccy", fontsize=14)
axes[0].set_ylabel("Count of Sensitivity Breaches", fontsize=12)
axes[0].tick_params(axis='x', rotation=45, labelsize=10)
axes[0].grid(True, linestyle='--', alpha=0.6)  # Grid background
add_bar_labels(axes[0])  # Add labels inside bars

# Tolerance Breach Chart
bars2 = axes[1].bar(
    tolerance_breach_counts.index,
    tolerance_breach_counts,
    color='lightcoral',
    edgecolor='black'  # Solid border
)
axes[1].set_title("Count of Tolerance Breaches by Product Type and Ccy", fontsize=14)
axes[1].set_ylabel("Count of Tolerance Breaches", fontsize=12)
axes[1].tick_params(axis='x', rotation=45, labelsize=10)
axes[1].grid(True, linestyle='--', alpha=0.6)  # Grid background
add_bar_labels(axes[1])  # Add labels inside bars

# New third chart (Immediate Attention Required)
# Filter for trades with both Sensitivity and Tolerance breaches
immediate_attention_df = df[
    (df['Sensitivity Breach'] == "TRUE") &
    (df['Tolerance Breach'] == True)
]
immediate_attention_counts = immediate_attention_df['Product_Ccy'].value_counts()

bars3 = axes[2].bar(
    immediate_attention_counts.index,
    immediate_attention_counts,
    color='darkred',  # Darker red to indicate urgency
    edgecolor='black'
)
axes[2].set_title("Breaks Requiring Immediate Attention (Both Sensitivity & Tolerance Breaches)", fontsize=14)
axes[2].set_ylabel("Count of Critical Breaches", fontsize=12)
axes[2].tick_params(axis='x', rotation=45, labelsize=10)
axes[2].grid(True, linestyle='--', alpha=0.6)
add_bar_labels(axes[2])

# Adjust layout for all three charts
plt.tight_layout()

# Second Chart: Trends in Sensitivity Breaches by Index and Curve Pillar

# Extract the year from Maturity Date and add a "Maturity Year" column
df['Maturity Year'] = pd.to_datetime(df['Maturity Date'], errors='coerce').dt.year.astype('Int64')

# Create a new column for the Index and Maturity Year concatenation
df['Index_Maturity'] = df['Index'] + "_" + df['Maturity Year'].astype(str)

# Filter data for the analysis
filtered_df = df[
    (df['Sensitivity Breach'] == "TRUE") &
    (df['Product Sub Type'] != "MTM Cross Currency Swap")
]

# Group by Index_Maturity and count Sensitivity Breaches
breach_counts = filtered_df.groupby(['Ccy', 'Index_Maturity']).size().reset_index(name='Breach Count')

# Get the top 5 Index Maturities **per currency**
top_5_per_currency = breach_counts.groupby('Ccy').apply(lambda x: x.nlargest(5, 'Breach Count')).reset_index(drop=True)

# Convert Final Source Load Time to datetime for plotting
df['Plot Date'] = pd.to_datetime(df['Final Source Load Time'], format='%d%m%Y')

# Create a figure with subplots for each unique currency
unique_ccys = top_5_per_currency['Ccy'].unique()
fig, axes = plt.subplots(len(unique_ccys), 1, figsize=(15, 5 * len(unique_ccys)), sharex=True)

# Ensure axes is iterable even if there's only one currency
if len(unique_ccys) == 1:
    axes = [axes]

# Plot data for each currency separately
for ax, ccy in zip(axes, unique_ccys):
    top_5_indices = top_5_per_currency[top_5_per_currency['Ccy'] == ccy]['Index_Maturity'].tolist()

    # Filter dataset for the selected currency and top 5 Index_Maturities
    ccy_data = df[
        (df['Ccy'] == ccy) &
        (df['Index_Maturity'].isin(top_5_indices)) &
        (df['Sensitivity Breach'] == "TRUE") &
        (df['Product Sub Type'] != "MTM Cross Currency Swap")
    ].groupby(['Plot Date', 'Index_Maturity']).size().unstack(fill_value=0)

    # Plot each Index Maturity in this currency's chart
    for column in ccy_data.columns:
        ax.plot(ccy_data.index, ccy_data[column], marker='o', label=column)  # Dots added back

    # Customize each subplot
    ax.set_title(f"Daily Sensitivity Breaches for {ccy} (Top 5 Index Maturities)", fontsize=14)
    ax.set_xlabel("Date", fontsize=12)
    ax.set_ylabel("Number of Breaches", fontsize=12)
    ax.grid(True, linestyle='--', alpha=0.6)
    ax.legend(title="Index Maturity", fontsize=8, bbox_to_anchor=(1.05, 1), loc='upper left')

# Rotate x-axis labels for readability
plt.xticks(rotation=45)

# Adjust layout to prevent overlap
plt.tight_layout()

# Show all plots
plt.show()




