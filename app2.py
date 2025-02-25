import pandas as pd
import glob
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import numpy as np
import streamlit as st
import re
from openpyxl.styles import Font


# Streamlit app title
st.title("IRS Report Processor + Tolerance Break Trend Analytics")

# File uploader for multiple Excel files
uploaded_files = st.file_uploader(
    "Drag and drop or select NAV reports (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True
)

if not uploaded_files:
    st.warning("Please upload at least one NAV report (.xlsx) to continue.")
    st.stop()

st.write(f"Uploaded {len(uploaded_files)} files for processing.")

# Function to extract valuation date from row 11 of 'IRS' sheet
def extract_valuation_date(file_path):
    """
    Extracts the Valuation Date from row 11 of the 'IRS' sheet in the Excel file.
    Converts the extracted date to DDMMYYYY format.
    """
    try:
        # Load workbook and explicitly select the "IRS" sheet
        wb = load_workbook(file_path, data_only=True)

        # Ensure "IRS" sheet exists before attempting to extract data
        if "IRS" not in wb.sheetnames:
            print(f"Warning: 'IRS' sheet not found in {file_path}")
            return None

        sheet = wb["IRS"]

        # Extract full string from row 11, column 1
        row_11_text = sheet.cell(row=11, column=1).value

        if row_11_text:
            # Look for "Valuation Date [DD-MMM-YYYY]"
            match = re.search(r'Valuation Date \[(\d{2})-(\w{3})-(\d{4})\]', row_11_text)
            if match:
                day, month_str, year = match.groups()

                # Convert month abbreviation to number
                month_map = {
                    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
                    "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
                }
                month = month_map.get(month_str, "00")

                return f"{day}{month}{year}"  # Return date in DDMMYYYY format

        return None  # Return None if no match is found

    except Exception as e:
        print(f"Error extracting valuation date from {file_path}: {e}")
        return None

# Initialize an empty list to store DataFrames
all_data = []

# Process each uploaded file, focusing only on the "IRS" sheet
for uploaded_file in uploaded_files:
    try:
        # Extract the Valuation Date from row 11
        valuation_date = extract_valuation_date(uploaded_file)

        # Read the "IRS" sheet
        raw_data = pd.read_excel(uploaded_file, sheet_name="IRS", header=[12, 13])

        # Generate column names by merging row 13 and 14
        new_columns = [
            f"{str(upper).strip()}_{str(lower).strip()}" if str(upper) != 'nan' else str(lower).strip()
            for upper, lower in zip(raw_data.columns.get_level_values(0), raw_data.columns.get_level_values(1))
        ]
        raw_data.columns = new_columns

        # Drop potential blank rows
        raw_data = raw_data.iloc[1:].reset_index(drop=True)

        # Add a Report Date column
        report_date = uploaded_file.name.split('-')[-1].split('.')[0] if '-' in uploaded_file.name else "Unknown Date"
        raw_data["Report Date"] = report_date

        # Add the Valuation Date column
        raw_data["Valuation Date"] = valuation_date

        # Append processed data
        all_data.append(raw_data)

    except Exception as e:
        st.error(f"Error processing {uploaded_file.name}: {e}")
        continue

# Combine all DataFrames
if all_data:
    data = pd.concat(all_data, ignore_index=True)
    st.success("All files processed successfully!")
else:
    st.error("No valid data found after processing.")
    st.stop()

# Filter the DataFrame to only keep relevant columns
filtered_columns = {
    "GTID_Unnamed: 0_level_1": "Trade ID 1",
    "Original GTID_Unnamed: 1_level_1": "Original GTID",
    "Counterparty / Clearing Member_MV Base": "Counterparty MV Base",
    "SS&C GlobeOp Source_MV Base": "SS&C MV Base",
    "Instrument Sub Type_Unnamed: 6_level_1": "Product Sub Type",
    "SS&C GlobeOp Source_IR DV01": "SS&C IR DV01",
    "SS&C GlobeOp Trade Attributes_Trade Date": "Trade Date",
    "SS&C GlobeOp Trade Attributes_Effective Date": "Effective Date",
    "SS&C GlobeOp Trade Attributes_Maturity Date": "Maturity Date",
    "SS&C GlobeOp Trade Attributes_Ccy": "Currency",
    "SS&C GlobeOp Trade Attributes_Notional": "Notional",
    "SS&C GlobeOp Trade Attributes_Rec Rate": "Rec Rate",
    "SS&C GlobeOp Trade Attributes_Pay Rate": "Pay Rate",
    "Counterparty/ Clearing Member_Final Source Load Time": "Final Source Load Time",
    "Final Source vs Counterparty / Clearing Member_Difference in MV": "Difference in MV",
    "Final Source vs Counterparty / Clearing Member_NAV Tolerance Analysis": "NAV Tolerance Analysis",
    "Final Source vs Counterparty / Clearing Member_Diff. in MV/IR DV01 or Diff. in MV/IDV01": "Diff. in MV/IR DV01",
    "Valuation Date": "Valuation Date",  # Add Valuation Date to filtered columns
    "Third Party_Name": "BBG Curve 4.30pm futures Snap",
    "Third Party2_Name": "LCH Curve w/ additional futs",
    "Third Party_MV Base": "BBG REFERENCE Curve MV (4.30 Futs Snap)",
    "Third Party2_MV Base": "LCH Test Curve MV"
}

# Ensure all columns in filtered_columns exist in the DataFrame
existing_columns = [col for col in filtered_columns.keys() if col in data.columns]
filtered_data = data[existing_columns].rename(columns=filtered_columns)

# Drop rows with missing Trade ID 1
filtered_data = filtered_data.dropna(subset=["Trade ID 1"])

# Convert Valuation Date to DDMMYYYY format
filtered_data['Valuation Date'] = pd.to_datetime(
    filtered_data['Valuation Date'], format='%d%m%Y', errors='coerce'
).dt.strftime('%d%m%Y')


# Set up new columns
filtered_data['Sensitivity Breach'] = None
filtered_data['Tolerance Breach'] = None

# Ask the user for the client name
client = st.text_input("Please enter the client you are analyzing (e.g., ASGARD): ").strip()

# Only proceed if the client is provided
if client:
    st.write(f"Processing data for client: **{client}**")

    if client.upper() == "ASGARD":
        # Apply conditions for Tolerance Breach
        filtered_data['Tolerance Breach'] = filtered_data['NAV Tolerance Analysis'].abs() > 1

    # Apply Sensitivity Breach logic to ASGARD
    ccy_list = ['USD', 'CAD', 'JPY', 'AUD', 'NZD', 'GBP', 'EUR', 'CHF', 'SEK', 'NOK']
    filtered_data['Sensitivity Breach'] = filtered_data.apply(
        lambda row: (
            "TRUE" if (
                row['Currency'] in ccy_list and
                (
                    (row['Product Sub Type'] == "Plain Vanilla" and abs(row['Diff. in MV/IR DV01']) > 2) or
                    (row['Product Sub Type'] == "OIS" and abs(row['Diff. in MV/IR DV01']) > 1) or
                    (row['Product Sub Type'] == "MTM Cross Currency Swap" and abs(row['Diff. in MV/IR DV01']) > 4)
                )
            ) else (
                "TRUE" if (
                    row['Currency'] not in ccy_list and
                    (
                        (row['Product Sub Type'] == "Plain Vanilla" and abs(row['Diff. in MV/IR DV01']) > 5) or
                        (row['Product Sub Type'] == "OIS" and abs(row['Diff. in MV/IR DV01']) > 8) or
                        (row['Product Sub Type'] == "MTM Cross Currency Swap" and abs(row['Diff. in MV/IR DV01']) > 9)
                    )
                ) else "FALSE"
            )
        ), axis=1
    )

    # Create index column
    filtered_data["Index"] = None
    filtered_data["Index_Maturity"] = None

    def get_index(row):
        def is_numeric_rate(value):
            if not isinstance(value, str):
                return True
            cleaned = value.replace('%', '').replace(' ', '')
            return cleaned.replace('.', '', 1).replace('-', '', 1).isdigit()

        rec_rate = row['Rec Rate']
        pay_rate = row['Pay Rate']

        if not is_numeric_rate(rec_rate):
            return rec_rate
        elif not is_numeric_rate(pay_rate):
            return pay_rate
        return None

    filtered_data['Index'] = filtered_data.apply(get_index, axis=1)

    # Extract the year from Maturity Date and add a "Maturity Year" column
    filtered_data['Maturity Year'] = pd.to_datetime(filtered_data['Maturity Date'], errors='coerce').dt.year.astype('Int64')

    # Create a new column for the Index and Maturity Year concatenation
    filtered_data['Index_Maturity'] = filtered_data['Index'] + "_" + filtered_data['Maturity Year'].astype(str)

    st.write(filtered_data)

    # Save the updated DataFrame to an Excel file
    output_file = "Processed_ASGARD_Report_with_Breaches.xlsx"
    filtered_data.to_excel(output_file, index=False, engine='openpyxl', sheet_name="Processed Report")

    # Apply conditional formatting for TRUE/FALSE
    wb = load_workbook(output_file)
    ws = wb["Processed Report"]
    filtered_data['Sensitivity Breach'] = filtered_data['Sensitivity Breach'].astype(str)
    filtered_data['Tolerance Breach'] = filtered_data['Tolerance Breach'].astype(str)

    columns_to_format = ["Sensitivity Breach", "Tolerance Breach"]
    true_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    false_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col in columns_to_format:
        if col in filtered_data.columns:
            col_idx = filtered_data.columns.get_loc(col) + 1
            for row in range(2, len(filtered_data) + 2):
                cell = ws.cell(row=row, column=col_idx)
                if str(cell.value).upper() == "TRUE":
                    cell.fill = true_fill
                elif str(cell.value).upper() == "FALSE":
                    cell.fill = false_fill
        else:
            st.warning(f"Warning: Column '{col}' not found in DataFrame. Skipping conditional formatting.")

    wb.save(output_file)
    st.write(f"Conditional formatting applied and saved to: {output_file}")

    # Ensure Tolerance Breach is a boolean before plotting
    filtered_data['Tolerance Breach'] = filtered_data['Tolerance Breach'].str.upper() == "TRUE"

    # Visualization - Breaches grouped by Product Type and Ccy
    filtered_data['Product_Ccy'] = filtered_data['Product Sub Type'] + "_" + filtered_data['Currency']

    sensitivity_breach_counts = filtered_data[filtered_data['Sensitivity Breach'] == "TRUE"]['Product_Ccy'].value_counts()
    tolerance_breach_counts = filtered_data[filtered_data['Tolerance Breach'] == True]['Product_Ccy'].value_counts()

    all_product_ccy = sensitivity_breach_counts.index.union(tolerance_breach_counts.index)
    sensitivity_breach_counts = sensitivity_breach_counts.reindex(all_product_ccy, fill_value=0)
    tolerance_breach_counts = tolerance_breach_counts.reindex(all_product_ccy, fill_value=0)

    # Plotting both charts stacked
    fig, axes = plt.subplots(3, 1, figsize=(14, 16), sharex=True)

    def add_bar_labels(ax):
        for bar in ax.patches:
            height = bar.get_height()
            if height > 0:
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    height * 0.5,
                    str(int(height)),
                    ha='center',
                    va='center',
                    fontsize=12,
                    fontweight='bold',
                    color='white'
                )

    # Sensitivity Breach Chart
    bars1 = axes[0].bar(sensitivity_breach_counts.index, sensitivity_breach_counts, color='skyblue', edgecolor='black')
    axes[0].set_title("Count of Sensitivity Breaches by Product Type and Ccy", fontsize=14)
    axes[0].set_ylabel("Count of Sensitivity Breaches", fontsize=12)
    axes[0].tick_params(axis='x', rotation=45, labelsize=10)
    axes[0].grid(True, linestyle='--', alpha=0.6)
    add_bar_labels(axes[0])

    # Tolerance Breach Chart
    bars2 = axes[1].bar(tolerance_breach_counts.index, tolerance_breach_counts, color='lightcoral', edgecolor='black')
    axes[1].set_title("Count of Tolerance Breaches by Product Type and Ccy", fontsize=14)
    axes[1].set_ylabel("Count of Tolerance Breaches", fontsize=12)
    axes[1].tick_params(axis='x', rotation=45, labelsize=10)
    axes[1].grid(True, linestyle='--', alpha=0.6)
    add_bar_labels(axes[1])

    # Immediate Attention Required Chart
    immediate_attention_df = filtered_data[
        (filtered_data['Sensitivity Breach'] == "TRUE") &
        (filtered_data['Tolerance Breach'] == True)
    ]
    immediate_attention_counts = immediate_attention_df['Product_Ccy'].value_counts()
    bars3 = axes[2].bar(immediate_attention_counts.index, immediate_attention_counts, color='darkred', edgecolor='black')
    axes[2].set_title("Breaks Requiring Immediate Attention (Both Sensitivity & Tolerance Breaches)", fontsize=14)
    axes[2].set_ylabel("Count of Critical Breaches", fontsize=12)
    axes[2].tick_params(axis='x', rotation=45, labelsize=10)
    axes[2].grid(True, linestyle='--', alpha=0.6)
    add_bar_labels(axes[2])

    plt.tight_layout()
    st.pyplot(fig)

    # Trend Analysis
    filtered_df = filtered_data[
        (filtered_data['Sensitivity Breach'] == "TRUE") &
        (filtered_data['Product Sub Type'] != "MTM Cross Currency Swap")
    ]
    breach_counts = filtered_df.groupby(['Currency', 'Index_Maturity']).size().reset_index(name='Breach Count')
    top_5_per_currency = breach_counts.groupby('Currency').apply(lambda x: x.nlargest(5, 'Breach Count')).reset_index(drop=True)

    filtered_data['Plot Date'] = pd.to_datetime(filtered_data['Valuation Date'], format='%d%m%Y')

    unique_ccys = top_5_per_currency['Currency'].unique()
    fig, axes = plt.subplots(len(unique_ccys), 1, figsize=(15, 5 * len(unique_ccys)), sharex=True)

    if len(unique_ccys) == 1:
        axes = [axes]

    for ax, ccy in zip(axes, unique_ccys):
        top_5_indices = top_5_per_currency[top_5_per_currency['Currency'] == ccy]['Index_Maturity'].tolist()
        ccy_data = filtered_data[
            (filtered_data['Currency'] == ccy) &
            (filtered_data['Index_Maturity'].isin(top_5_indices)) &
            (filtered_data['Sensitivity Breach'] == "TRUE") &
            (filtered_data['Product Sub Type'] != "MTM Cross Currency Swap")
        ].groupby(['Plot Date', 'Index_Maturity']).size().unstack(fill_value=0)

        for column in ccy_data.columns:
            ax.plot(ccy_data.index, ccy_data[column], marker='o', label=column)

        ax.set_title(f"Daily Sensitivity Breaches for {ccy} (Top 5 Index Maturities)", fontsize=14)
        ax.set_xlabel("Date", fontsize=12)
        ax.set_ylabel("Number of Breaches", fontsize=12)
        ax.grid(True, linestyle='--', alpha=0.6)
        ax.legend(title="Index Maturity", fontsize=8, bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.xticks(rotation=45)
    plt.tight_layout()
    st.pyplot(fig)

    # Filter the data for Sensitivity Breaches and exclude MTM Cross Currency Swap
    filtered_df = filtered_data[
        (filtered_data['Sensitivity Breach'] == "TRUE") &
        (filtered_data['Product Sub Type'] != "MTM Cross Currency Swap")
        ]

    # Group by Currency and Index_Maturity to get breach counts
    breach_counts = filtered_df.groupby(['Currency', 'Index_Maturity']).size().reset_index(name='Breach Count')

    # Get the top 5 Index Maturities per Currency
    top_5_per_currency = breach_counts.groupby('Currency').apply(lambda x: x.nlargest(5, 'Breach Count')).reset_index(
        drop=True)

    # Debug: Print top_5_per_currency to verify data
    st.write("Top 5 Index Maturities per Currency:")
    st.write(top_5_per_currency)

    # Get unique currencies
    unique_currencies = top_5_per_currency['Currency'].unique()

    # Create a separate bubble chart for each currency
    for currency in unique_currencies:
        # Filter data for the current currency
        currency_data = top_5_per_currency[top_5_per_currency['Currency'] == currency]

        # Create a new figure for the current currency
        plt.figure(figsize=(10, 6))

        # Plot the bubble chart
        plt.scatter(
            currency_data['Index_Maturity'],  # X-axis: Index Maturity
            [1] * len(currency_data),  # Y-axis: Dummy value (all points on the same line)
            s=currency_data['Breach Count'] * 100,  # Bubble size: Breach Count (scaled for visibility)
            alpha=0.6,  # Transparency
            color='skyblue',  # Bubble color
            edgecolor='black'  # Bubble edge color
        )

        # Add labels and title
        plt.title(f"Sensitivity Breaches for {currency} (Top 5 Index Maturities)", fontsize=16)
        plt.xlabel("Index Maturity", fontsize=14)
        plt.ylabel("", fontsize=14)  # No label for Y-axis
        plt.grid(True, linestyle='--', alpha=0.6)

        # Rotate x-axis labels for better readability
        plt.xticks(rotation=45)

        # Display the chart in Streamlit
        st.pyplot(plt)

        # Clear the figure to avoid overlapping plots
        plt.clf()

    # Provide download link for the processed Excel file
    st.subheader("Download Processed Data")
    st.write("Click the button below to download the processed Excel file.")
    with open(output_file, "rb") as file:
        btn = st.download_button(
            label="Download Excel",
            data=file,
            file_name=output_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    ### Comparative Analysis

    # Dropdown for Index
    index_options = filtered_df['Index'].unique()
    selected_index = st.selectbox("Select Index", index_options)

    # Dropdown for Product Sub Type
    product_sub_type_options = filtered_df['Product Sub Type'].unique()
    selected_product_sub_type = st.selectbox("Select Product Sub Type", product_sub_type_options)

    # Text input for Trade IDs
    trade_ids_input = st.text_input("Enter Trade IDs (comma-separated)")

    # Split the input into a list of Trade IDs
    selected_trade_ids = [tid.strip() for tid in trade_ids_input.split(",")] if trade_ids_input else []

    # Filter the data based on user selections
    filtered_df2 = filtered_df[
        (filtered_df['Product Sub Type'] == selected_product_sub_type) &
        (filtered_df['Index'] == selected_index) &
        (filtered_df['Trade ID 1'].isin(selected_trade_ids))
        ]

    # Convert Valuation Date to datetime for proper time series plotting
    filtered_df2['Valuation Date'] = pd.to_datetime(filtered_df2['Valuation Date'], format='%d%m%Y')

    ## Perform comparative analysis

    # Ensure columns are numeric
    filtered_df2['BBG REFERENCE Curve MV (4.30 Futs Snap)'] = filtered_df2[
        'BBG REFERENCE Curve MV (4.30 Futs Snap)'].replace({'TRUE': 1, 'FALSE': 0}).astype(float)
    filtered_df2['LCH Test Curve MV'] = pd.to_numeric(filtered_df2['LCH Test Curve MV'], errors='coerce')
    filtered_df2['BBG REFERENCE Curve MV (4.30 Futs Snap)'] = pd.to_numeric(
        filtered_df2['BBG REFERENCE Curve MV (4.30 Futs Snap)'], errors='coerce')

    # Ensure denominator (SS&C IR DV01) is non-zero to prevent division errors
    filtered_df2['SS&C IR DV01'] = filtered_df2['SS&C IR DV01'].replace(0, np.nan)

    # Calculate new basis points differences
    filtered_df2['BBG REFERENCE Curve MV Diff'] = (
            (filtered_df2['BBG REFERENCE Curve MV (4.30 Futs Snap)'] - filtered_df2['Counterparty MV Base']) /
            filtered_df2['SS&C IR DV01']
    )

    filtered_df2['LCH Test Curve MV Diff'] = (
            (filtered_df2['LCH Test Curve MV'] - filtered_df2['Counterparty MV Base']) /
            filtered_df2['SS&C IR DV01']
    )

    # Ensure the new columns are numeric
    filtered_df2['LCH Test Curve MV Diff'] = pd.to_numeric(filtered_df2['LCH Test Curve MV Diff'],
                                                           errors='coerce').fillna(0)

    # Group by Valuation Date and calculate averages
    time_series_data = filtered_df2.groupby('Valuation Date').agg({
        'Diff. in MV/IR DV01': 'mean',  # Average Sensitivity Breaches
        'LCH Test Curve MV Diff': 'mean',  # Average LCH Curve
        'BBG REFERENCE Curve MV Diff': 'mean'  # Average BBG Curve
    }).reset_index()

    # Create the time series plot
    fig, ax = plt.subplots(figsize=(14, 8))

    # Plot Sensitivity Breaches
    ax.plot(
        time_series_data['Valuation Date'],
        time_series_data['Diff. in MV/IR DV01'],
        label='Average BBG MODEL Sensitivity Breaches',
        marker='o',
        linestyle='-',
        color='blue'
    )

    # Plot LCH Curve vs CPTY Diff. in MV/IR DV01
    ax.plot(
        time_series_data['Valuation Date'],
        time_series_data['LCH Test Curve MV Diff'],
        label='LCH Test Curve MV Diff',
        marker='s',
        linestyle='-',
        color='green'
    )

    # Plot BBG Curve 4.30pm futs snap vs CPTY Diff. in MV/IR DV01
    ax.plot(
        time_series_data['Valuation Date'],
        time_series_data['BBG REFERENCE Curve MV Diff'],
        label='BBG REFERENCE Curve MV Diff',
        marker='^',
        linestyle='-',
        color='red'
    )

    # Add labels and title
    ax.set_title("Time Series of Average Metrics", fontsize=16)
    ax.set_xlabel("Valuation Date", fontsize=14)
    ax.set_ylabel("Average Value", fontsize=14)
    ax.grid(True, linestyle='--', alpha=0.6)

    # Add legend and ensure it's fully visible
    ax.legend(title="Metrics", loc="upper left", bbox_to_anchor=(1.02, 1), borderaxespad=0.)

    # Rotate x-axis labels for better readability
    plt.xticks(rotation=45)

    # Adjust layout to prevent clipping
    plt.tight_layout()

    # Display the plot in Streamlit
    st.pyplot(fig)

    ## Add Clustered Column Chart for Differences > 2

    # Flagging values > 2
    filtered_df2['Diff > 2 (BBG MODEL)'] = (filtered_df2['Diff. in MV/IR DV01'] > 2).astype(int)
    filtered_df2['Diff > 2 (LCH Curve)'] = (filtered_df2['LCH Test Curve MV Diff'] > 2).astype(int)
    filtered_df2['Diff > 2 (BBG Curve)'] = (filtered_df2['BBG REFERENCE Curve MV Diff'] > 2).astype(int)

    # Grouping by valuation date to count occurrences
    count_diff_df = filtered_df2.groupby('Valuation Date').agg({
        'Diff > 2 (BBG MODEL)': 'sum',
        'Diff > 2 (LCH Curve)': 'sum',
        'Diff > 2 (BBG Curve)': 'sum'
    }).reset_index()

    # Convert Valuation Date to datetime and then to DDMMYYYY format
    count_diff_df['Valuation Date'] = pd.to_datetime(count_diff_df['Valuation Date']).dt.strftime('%d%m%Y')

    # Create a clustered column chart
    fig2, ax2 = plt.subplots(figsize=(14, 8))

    # Define bar width and x positions
    bar_width = 0.25
    x = np.arange(len(count_diff_df['Valuation Date']))

    # Define softer shades for the columns
    soft_blue = '#A6CEE3'  # Soft blue
    soft_green = '#B2DF8A'  # Soft green
    soft_red = '#FB9A99'  # Soft red

    # Plot bars separately for each metric
    ax2.bar(
        x - bar_width, count_diff_df['Diff > 2 (BBG MODEL)'],
        width=bar_width,
        color=soft_blue,
        edgecolor='black',  # Thick outline
        linewidth=2,  # Outline thickness
        label='BBG MODEL >2'
    )
    ax2.bar(
        x, count_diff_df['Diff > 2 (LCH Curve)'],
        width=bar_width,
        color=soft_green,
        edgecolor='black',  # Thick outline
        linewidth=2,  # Outline thickness
        label='LCH Curve >2'
    )
    ax2.bar(
        x + bar_width, count_diff_df['Diff > 2 (BBG Curve)'],
        width=bar_width,
        color=soft_red,
        edgecolor='black',  # Thick outline
        linewidth=2,  # Outline thickness
        label='BBG Curve >2'
    )

    # Formatting the chart
    ax2.set_xticks(x)
    ax2.set_xticklabels(count_diff_df['Valuation Date'], rotation=45)
    ax2.set_title("Count of Differences > 2 Over Time (Clustered Columns)", fontsize=16)
    ax2.set_xlabel("Valuation Date", fontsize=14)
    ax2.set_ylabel("Count of Differences > 2", fontsize=14)

    # Add grid background
    ax2.grid(True, linestyle='--', alpha=0.6)

    # Add legend and ensure it's fully visible
    ax2.legend(title="Metrics", loc="upper right")

    # Improve layout
    plt.tight_layout()

    # Display the plot in Streamlit
    st.pyplot(fig2)

# Add a new column to flag exceptions
    filtered_data['Exception Flag'] = filtered_data[['Counterparty MV Base', 'SS&C MV Base']].isnull().any(axis=1)


# Create a DataFrame for exceptions
    exceptions_df = filtered_data[filtered_data['Exception Flag']]



# Select the relevant columns for the exceptions report
    exceptions_report_columns = [
        'Valuation Date', 'Trade ID 1', 'Product Sub Type', 'Trade Date',
        'Maturity Date', 'Currency', 'Notional', 'Pay Rate', 'Rec Rate', 'SS&C IR DV01',
        'Counterparty MV Base', 'SS&C MV Base'
    ]

    exceptions_report = exceptions_df[exceptions_report_columns].sort_values(by=["Valuation Date", "Trade ID 1"])



# Save the exceptions report to an Excel file
    exceptions_output_file = "Exceptions_Report.xlsx"
    exceptions_report.to_excel(exceptions_output_file, index=False, engine='openpyxl', sheet_name="Exceptions Report")

    wb = load_workbook(exceptions_output_file)
    ws = wb["Exceptions Report"]

    # Get all unique Valuation Dates
    valuation_dates = exceptions_report["Valuation Date"].unique()

    # Apply bold styling for headers
    bold_font = Font(bold=True)

    row_offset = 2  # Start after header
    for val_date in valuation_dates:
        ws.cell(row=row_offset, column=1, value=str(val_date)).font = bold_font
        row_offset += len(exceptions_report[exceptions_report["Valuation Date"] == val_date]) + 1  # Leave a blank row

    # Save workbook
    wb.save(exceptions_output_file)

    # Apply conditional formatting to highlight missing values
    exception_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in ['Counterparty MV Base', 'SS&C MV Base']:
        col_idx = exceptions_report.columns.get_loc(col) + 1  # Get correct column index
        for row in range(2, len(exceptions_report) + 2):  # Start from row 2
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None or cell.value == "":  # Check for empty values
                cell.fill = exception_fill

    wb.save(exceptions_output_file)



# Provide download link for the exceptions Excel file
    st.subheader("Download Exceptions Report")
    st.write("Click the button below to download the exceptions Excel file.")
    with open(exceptions_output_file, "rb") as file:
        btn = st.download_button(
            label="Download Exceptions Excel",
            data=file,
            file_name=exceptions_output_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )